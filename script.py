import openpyxl
wb = openpyxl.load_workbook('sorter.xlsx')
ws = wb.active


with open('swaps.txt') as f:
    myLines = f.readlines()
r = 2
myLines = [line.rstrip('\n') for line in myLines]
for line in myLines:
    ws.cell(row = r, column = 2).value = int(line)
    r+=1


with open('comp.txt') as f1:
    myLines1 = f1.readlines()
r1 = 2
myLines1 = [line1.rstrip('\n') for line1 in myLines1]
for line1 in myLines1:
    ws.cell(row = r1, column = 3).value = int(line1)
    r1+=1
   


with open('numbers.txt') as f2:
    myLines2 = f2.readlines()
r2 = 2
myLines2 = [line2.rstrip('\n') for line2 in myLines2]
for line2 in myLines2:
    ws.cell(row = r2, column = 1).value = int(line2)
    r2+=1
wb.save('sorter.xlsx')